# -*- coding:utf-8 -*-
#!/usr/bin/env python3

from xml.dom.minidom import parse
import xml.dom.minidom
import os
#没有openpyxl模块的需要安装，可通过pip install openpyxl指令安装
from openpyxl import Workbook,load_workbook

#coding=utf-8
import sys

class ScanCueXml(object):
   def __init__(self):
      self.targetExcelPath = os.path.join(os.path.dirname(__file__), "usingCueNode.xlsx")
      #所有cue节点type字典
      self.allCueNodeTypeDir = {}
      #正在使用cue节点字典key：cueType  value:(cueType,cueName)
      self.cueDir = {}
      #cue与使用了该cue的Graph信息。key:cueType Value:(graphName)
      self.curNodeGraphDir = {}
      #需要扩展添加的属性值列表
      self.toAddDataKeyList = ("Type","Name")
      #未使用的节点表里所需要统计的字段
      self.unUsedCueAttrList = ["CueType","Name","Desc"]

   def __del__(self):
      # print("destroy")
      del self.allCueNodeTypeDir
      del self.cueDir
      del self.curNodeGraphDir
      del self.toAddDataKeyList
      del self.unUsedCueAttrList

   def onError(self,OSError):
      print("遍历文件目录错误 :{}".format(OSError))

   def codeTransfer(self,_value):
      value = _value
      if isinstance(value,unicode):
         value = value.encode("unicode-escape")
      else:
         #从设置的默认编码utf-8转到中转unicode后再用目标编码unicode-escape输出
         value = value.decode("UTF-8")
         value = value.encode("unicode-escape")
      value = str.strip(str(value))
      return value

   def toUnicode(self,_value):
      # value = str(value[1]).replace('u\'','\'')
      if _value is None:
         return "为空"
      return _value.decode("unicode-escape")

   def getSubElementData(self,parentElement):
      dataList = []
      for value in self.toAddDataKeyList:
         tsubElement = parentElement.getElementsByTagName(value) [0] if parentElement.getElementsByTagName(value).length > 0 else None
         if tsubElement is not None:
            # print ("{}:{} {}".format("Name",tsubElement,tsubElement.childNodes[0].data))
            value = self.codeTransfer(tsubElement.childNodes[0].data)
            dataList.append(value)
      if len(dataList) > 0:#self.cueDir.has_key(dataList[0])
         self.cueDir.update({str.strip(dataList[0]) : dataList})

   def fillSheetCell(self,_value,addCol = True,addRow = False,_row = 0,_col = 0):
      if self.workSheet is None:
         return

      self.workSheet.cell(row=self.curRow, column=self.curCol, value=_value)
      if addCol:
         self.curCol = self.curCol + 1

      if addRow:
         self.curRow = self.curRow + 1
         #换行重置列数
         self.curCol = 1

   def recordUnusedCueData(self,_unusedCueList,_excelWorkBook):
      sheet = _excelWorkBook.create_sheet("unUsedCueType",1)
      row = 1
      col = 1

      #unUsedCueType sheet attr,添加字段名
      count = 1
      for _value in self.unUsedCueAttrList:
         sheet.cell(row=row, column=count, value = _value)
         count += 1
      row += 1

      #添加字段值
      keyList = []
      for key,value in _unusedCueList.items():
         keyList.append(int(str.strip(key)))
      keyList.sort()
      for key in keyList:
         valueList =  _unusedCueList.get(str(key),None)#"无此key:{}".format(key)
         if valueList is not None:
            count = len(valueList)
            for index in range(count):
               sheet.cell(row=row, column=col + index, value = self.toUnicode(valueList[index]))
            row += 1

   def initAllCueData(self):
      #方式一：读策划整理的表格
      # path = os.path.abspath(os.path.join(os.path.dirname(__file__),"cueType.xlsx"))
      # workSheet = load_workbook(path).active
      # count = 0
      # for valueTuple in workSheet.values:
      #    if count == 0:
      #       count += 1
      #       continue
      #    count += 1
      #    cuetype = valueTuple[0]
      #    cueName = valueTuple[1]
      #    cueDesc = valueTuple[2]
      #方式二：读cue.xml
      path = os.path.abspath(os.path.join(os.path.dirname(__file__),"../../engine_trunk/tools/CharacterEditor-Win64/res"))
      dirData = os.walk(path)
      for path,dir_name_list,file_name_list in dirData:
         for file_name in file_name_list:
            if file_name == "cue.xml":
               print("start parse cue.xml")
               subFilePath = os.path.join(path, file_name)
               DOMTree = xml.dom.minidom.parse(subFilePath)
               collection = DOMTree.documentElement

               #现在的cue.xml定义的cue节点元素不是统一的，得从父节点处操作，如果后续添加新cue节点时改了规则，这里处理可能会异常
               for element in collection.getElementsByTagName("CueType"):
                  dataList = []
                  # self.getAllSubElementData(element)
                  parentNode = element.parentNode
                  for node in parentNode.childNodes:
                     for attr in self.unUsedCueAttrList:
                        if node.nodeName == attr:
                           dataList.append(self.codeTransfer(node.childNodes[0].data))
                  #key:cueType : value:[attrList's attr]
                  self.allCueNodeTypeDir.update({self.codeTransfer(element.childNodes[0].data) : dataList})


   def closePreOpenExcel(self):
      temp_file = os.path.join(os.path.dirname(__file__), "~$usingCueNode.xlsx")
      if os.path.exists(temp_file):
         print('ERROR:: usingCueNodeexcel is opened,please close it firstly')
         # load_workbook(self.targetExcelPath).close()
         return False
      else:
         return True
         
   #记录每个cue所关联的graph信息
   def mapGraphToCue(self,_parentElement,_graphDataList):
      tsubElement = _parentElement.getElementsByTagName("Type") [0] if _parentElement.getElementsByTagName("Type").length > 0 else None
      cueType = None
      if tsubElement is not None:
         # print ("{}:{} {}".format("Name",tsubElement,tsubElement.childNodes[0].data))
         cueType = self.codeTransfer(tsubElement.childNodes[0].data)
      if cueType is None:
         print("ERROR::_parentElement have no cue")
         return

      if not self.curNodeGraphDir.has_key(cueType):
         self.curNodeGraphDir.update({ cueType:{} })#一个graph可能使用多个相同的cue，用字典来存
      
      preDataDir = self.curNodeGraphDir[cueType]
      for value in _graphDataList:
         if not preDataDir.has_key(value):
            preDataDir.update( {value : value} )

   def scan(self):
      if not self.closePreOpenExcel():
         return

      self.initAllCueData()
      # self.cuePath = "D:/G109/Project/engine_trunk/tools/CharacterEditor-Win64/res/cue.xml"
      excelWorkBook = Workbook()
      #当前选中的worksheet
      self.workSheet = excelWorkBook.active
      self.workSheet.title = "usedCue"

      #列字段依次为：graph路径，graph名，用到的cue节点的Name     #_TrackName，
      self.curRow = 1
      self.curCol = 1
      self.fillSheetCell("graph name")
      # self.fillSheetCell("graph path")
      # self.fillSheetCell("cue node type/name")
      count = 0
      for value in self.toAddDataKeyList:
         attr = "cur node {}".format(value)
         addRow = False
         addCol = True
         if count >= (len(self.toAddDataKeyList) - 1):
            addRow = True
         self.fillSheetCell(attr,addCol,addRow)
         count += 1

      #获取当前根路径
      # graphRootPath = "D:/G109/Project/code_trunk/src/Package/Graph"
      # print(os.path.abspath(os.path.join(os.path.dirname(__file__),"../../code_trunk/src/Package/Graph")))
      graphRootPath = os.path.abspath(os.path.join(os.path.dirname(__file__),"../../code_trunk/src/Package/Graph"))
      
      print("=============================== start scan ===============================")
      dirData = os.walk(graphRootPath)#,topdown = True,onerror = self.onError,followlinks = False
      for path,dir_name_list,file_name_list in dirData:
         for dir_name in dir_name_list:
            dirPath = os.path.join(path, dir_name)
            # print("graph目录:{}，路径:{}".format(dir_name,dirPath))
            # self.fillSheetCell(dirPath)
            # self.fillSheetCell(dir_name)
            subDirData = os.walk(dirPath)
            for subpath,subdir_name_list,subfile_name_list in subDirData:
               for subfile_name in subfile_name_list: 
                  if os.path.splitext(subfile_name)[-1] == ".graph":
                     subFilePath = os.path.join(subpath, subfile_name)
                     # print("{}目录下的graph文件:{}，路径:{}".format(dir_name,subfile_name,subFilePath))

                     #需要跟cueType关联的graph数据
                     graphDataList = [subfile_name]#subFilePath

                     # 使用minidom解析器打开 XML 文档
                     DOMTree = xml.dom.minidom.parse(subFilePath)

                     collection = DOMTree.documentElement
                     for element in collection.getElementsByTagName("Cue"):
                        self.getSubElementData(element)
                        self.mapGraphToCue(element,graphDataList)
      
      # print(os.path.join(os.getcwd(), "cleanUpCueNode.xlsx"))
      keyList = []
      valueList = []
      for key,value in self.cueDir.items():
         keyList.append(int(key))
      
      if len(keyList) > 0 :
         keyList.sort()

      for key in keyList :
         key = str(key)
         #cue相关的graph信息填充
         graphDir = self.curNodeGraphDir[key]
         graphDataStr = ""
         for graphDirKey,graphDirValue in graphDir.items():
            graphDataStr = "{},{}".format(graphDataStr,self.toUnicode(graphDirValue))
         self.fillSheetCell(graphDataStr[1:])

         #cue信息填充
         value = self.cueDir[key]
         self.fillSheetCell(self.toUnicode(key))
         # value = str(value[1]).replace('u\'','\'')
         value = self.toUnicode(value[1])
         self.fillSheetCell(value,True,True)

      unusedCueDir = {}
      for key,value in self.allCueNodeTypeDir.items():
         #记录未使用到的cue
         if not self.cueDir.has_key(key):
            unusedCueDir.update({key:value})

      self.recordUnusedCueData(unusedCueDir,excelWorkBook)

      excelWorkBook.save(self.targetExcelPath)
      print("=============================== end scan ===============================")
      
      
ScanCueXml().scan()
#python执行完后停在解释器窗口
os.system("pause")